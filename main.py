from tkinter import Tk, Label, Button, filedialog, messagebox, StringVar, Frame, Text, Scrollbar, RIGHT, Y, END
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from tkinter import Tk, Label, Button, filedialog, messagebox, StringVar, Frame

# Diccionario de mapeo de respuestas a puntuaciones según tipo de pregunta
puntuaciones = {
    'negativas': {
        'Siempre': 4,
        'Casi siempre': 3,
        'Algunas veces': 2,
        'Casi nunca': 1,
        'Nunca': 0,
        'Casi nuca': 1  # Posible error de dedo, pero se mantiene por si aparece en los datos
    },
    'positivas': {
        'Siempre': 0,
        'Casi siempre': 1,
        'Algunas veces': 2,
        'Casi nunca': 3,
        'Nunca': 4
    }
}

# Diccionario de categorías y subcategorías según el cuestionario
categorias = {
    'Ambiente de trabajo': [1, 2, 3],
    'Factores propios de la actividad': {
        'Carga de trabajo': [4, 5, 6, 7, 8, 9, 41, 42, 43],
        'Cargas de alta responsabilidad': [10, 11],
        'Cargas contradictorias o inconsistentes': [12, 13],
        'Falta de control sobre el trabajo': [20, 21, 22, 18, 19, 26, 27],
    },
    'Organización del tiempo de trabajo': {
        'Jornada de trabajo': [14, 15],
        'Interferencia en la relación trabajo-familia': [16, 17]
    },
    'Liderazgo y relaciones en el trabajo': {
        'Liderazgo': [23, 24, 25, 28, 29],
        'Relaciones en el trabajo': [30, 31, 32, 33],
        'Violencia': [34, 35, 36, 37, 38, 39, 40],
        'Deficiente relación con los colaboradores que supervisa': [44, 45, 46]
    }
}


def determinar_nivel_riesgo(puntuacion):
    """
    Determina el nivel de riesgo psicosocial según la puntuación total.
    """
    if puntuacion < 20:
        return "Nulo o despreciable"
    elif 20 <= puntuacion < 45:
        return "Bajo"
    elif 45 <= puntuacion < 70:
        return "Medio"
    elif 70 <= puntuacion < 90:
        return "Alto"
    else:
        return "Muy alto"


def calcular_puntuaciones(df):
    """
    Calcula las puntuaciones totales y por categoría para cada trabajador.
    Devuelve dos DataFrames: uno con los resultados generales y otro con los detalles por pregunta.
    """
    resultados = []
    detalles_por_pregunta = []

    for _, row in df.iterrows():
        puntuacion_total = 0
        detalles = {'Nombre': row['Nombre Completo del trabajador']}
        detalles_preguntas = {'Nombre': row['Nombre Completo del trabajador']}

        # Recorre todas las preguntas (1 a 46)
        for i in range(1, 47):
            col_name = f"{i}" if i < 0 else f"{i}"
            if col_name in row.index:
                respuesta = row[col_name]
                # Si la respuesta está vacía o es NaN, tratar como "Nunca"
                if pd.isna(respuesta) or respuesta == "":
                    respuesta = "Nunca"
                tipo = 'negativas' if (
                    1 <= i <= 17 or 34 <= i <= 46) else 'positivas'
                puntuacion = puntuaciones[tipo].get(respuesta, 0)
                puntuacion_total += puntuacion
                detalles[f"P{i}"] = puntuacion
                # Guardar la respuesta original (ya como "Nunca" si estaba vacía)
                detalles_preguntas[f"P{i}"] = respuesta

        nivel_riesgo = determinar_nivel_riesgo(puntuacion_total)

        # Calcular puntuaciones por categoría y subcategoría
        categorias_puntuacion = {}
        cat = 'Ambiente de trabajo'
        preguntas = categorias[cat]
        punt_cat = sum(detalles.get(f"P{p}", 0) for p in preguntas)
        categorias_puntuacion[cat] = punt_cat

        for cat, subcats in [k for k in categorias.items() if isinstance(k[1], dict)]:
            punt_cat_total = 0
            for subcat, preguntas in subcats.items():
                punt_subcat = sum(detalles.get(f"P{p}", 0) for p in preguntas)
                categorias_puntuacion[f"{cat} - {subcat}"] = punt_subcat
                punt_cat_total += punt_subcat
            categorias_puntuacion[cat] = punt_cat_total

        resultados.append({
            'Nombre': detalles['Nombre'],
            'Puntuación Total': puntuacion_total,
            'Nivel de Riesgo': nivel_riesgo,
            **categorias_puntuacion
        })

        detalles_por_pregunta.append(detalles_preguntas)

    return pd.DataFrame(resultados), pd.DataFrame(detalles_por_pregunta)


def generar_recomendaciones(nivel):
    """
    Devuelve una recomendación según el nivel de riesgo.
    """
    recomendaciones = {
        "Nulo o despreciable": "El riesgo resulta despreciable por lo que no se requiere medidas adicionales.",
        "Bajo": "Es necesario una mayor difusión de la política de prevención de riesgos psicosociales y programas para: la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral.",
        "Medio": "Se requiere revisar la política de prevención de riesgos psicosociales y programas para la prevención de los factores de riesgo psicosocial, la promoción de un entorno organizacional favorable y la prevención de la violencia laboral, así como reforzar su aplicación y difusión, mediante un Programa de intervención.",
        "Alto": "Se requiere realizar un análisis de cada categoría y dominio, de manera que se puedan determinar las acciones de intervención apropiadas a través de un Programa de intervención...",
        "Muy alto": "Se requiere realizar el análisis de cada categoría y dominio para establecer las acciones de intervención apropiadas, mediante un Programa de intervención que deberá incluir evaluaciones específicas..."
    }
    return recomendaciones.get(nivel, "Nivel de riesgo no reconocido.")


def crear_reporte_individual(row, detalles_preguntas, area_adscrita):
    """
    Crea un archivo Excel con el reporte individual de un trabajador.
    """
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Individual"

    # Estilos para el formato del reporte
    bold_font = Font(bold=True)
    center_alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4F81BD",
                              end_color="4F81BD", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    # Color según nivel de riesgo
    color_nivel = {
        "Nulo o despreciable": "C6EFCE",
        "Bajo": "D9EAD3",
        "Medio": "FFF2CC",
        "Alto": "FCE5CD",
        "Muy alto": "F4CCCC"
    }.get(row['Nivel de Riesgo'], "FFFFFF")

    # Encabezado del reporte
    mes_actual = datetime.now().strftime("%B %Y").upper()
    ws.merge_cells('A1:G1')
    ws['A1'] = f"RESULTADOS DE EVALUACIÓN DE RIESGOS PSICOSOCIALES ({mes_actual})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    ws['A3'] = "Trabajador"
    ws['B3'] = row['Nombre']
    ws['A4'] = "Área adscrita"
    ws['B4'] = area_adscrita
    ws['A5'] = "Nivel de riesgo"
    ws['B5'] = row['Nivel de Riesgo']
    ws['B5'].fill = PatternFill(
        start_color=color_nivel, end_color=color_nivel, fill_type="solid")

    # Encabezado de la tabla de resultados
    encabezados = [
        "Categoría", "Dominio", "Dimensión",
        "Puntuación de dimensión",
        "Resultado del cuestionario",
        "Calificación de la categoría",
        "Resultado por dominio"
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=7, column=col, value=encabezado)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Mapeo de preguntas a dimensiones (para mostrar resultados por dimensión)
    mapeo_dimensiones = {
        # Ambiente de trabajo
        "Condiciones peligrosas e inseguras": [1],
        "Condiciones deficientes e insalubres": [2],
        "Trabajos peligrosos": [3],
        # Factores propios de la actividad - Carga de trabajo
        "Cargas cuantitativas": [4, 5],
        "Ritmos de trabajo acelerado": [6],
        "Carga mental": [7, 8, 9],
        "Cargas psicológicas emocionales": [41, 42, 43],
        "Cargas de alta responsabilidad": [10, 11],
        "Cargas contradictorias o inconsistentes": [12, 13],
        # Factores propios de la actividad - Falta de control
        "Falta de control y autonomía sobre el trabajo": [20, 21, 22],
        "Limitada o nula posibilidad de desarrollo": [18, 19],
        "Limitada o inexistente capacitación": [26, 27],
        # Organización del tiempo de trabajo - Jornada
        "Jornadas de trabajo extensas": [14, 15],
        # Organización del tiempo de trabajo - Interferencia
        "Influencia del trabajo fuera del centro laboral": [16],
        "Influencia de las responsabilidades familiares": [17],
        # Liderazgo y relaciones - Liderazgo
        "Escasa claridad de funciones": [23, 24, 25],
        "Características del liderazgo": [28, 29],
        # Liderazgo y relaciones - Relaciones
        "Relaciones sociales en el trabajo": [30, 31, 32, 33],
        "Deficiente relación con los colaboradores que supervisa": [44, 45, 46],
        # Liderazgo y relaciones - Violencia
        "Violencia laboral": [34, 35, 36, 37, 38, 39, 40]
    }

    # Datos de las categorías y dimensiones (para mostrar en la tabla)
    categorias_data = [
        # Ambiente de trabajo
        ["Ambiente de trabajo", "Condiciones en el ambiente de trabajo",
            "Condiciones peligrosas e inseguras"],
        ["", "", "Condiciones deficientes e insalubres"],
        ["", "", "Trabajos peligrosos"],

        # Factores propios de la actividad - Carga de trabajo
        ["Factores propios de la actividad",
            "Carga de trabajo", "Cargas cuantitativas"],
        ["", "", "Ritmos de trabajo acelerado"],
        ["", "", "Carga mental"],
        ["", "", "Cargas psicológicas emocionales"],
        ["", "Cargas de alta responsabilidad", "Cargas de alta responsabilidad"],
        ["", "Cargas contradictorias o inconsistentes",
            "Cargas contradictorias o inconsistentes"],

        # Factores propios de la actividad - Falta de control
        ["", "Falta de control sobre el trabajo",
            "Falta de control y autonomía sobre el trabajo"],
        ["", "", "Limitada o nula posibilidad de desarrollo"],
        ["", "", "Limitada o inexistente capacitación"],

        # Organización del tiempo de trabajo - Jornada
        ["Organización del tiempo de trabajo",
            "Jornada de trabajo", "Jornadas de trabajo extensas"],

        # Organización del tiempo de trabajo - Interferencia
        ["", "Interferencia en la relación trabajo-familia",
            "Influencia del trabajo fuera del centro laboral"],
        ["", "", "Influencia de las responsabilidades familiares"],

        # Liderazgo y relaciones - Liderazgo
        ["Liderazgo y relaciones en el trabajo",
            "Liderazgo", "Escasa claridad de funciones"],
        ["", "", "Características del liderazgo"],

        # Liderazgo y relaciones - Relaciones
        ["", "Relaciones en el trabajo", "Relaciones sociales en el trabajo"],
        ["", "", "Deficiente relación con los colaboradores que supervisa"],

        # Liderazgo y relaciones - Violencia
        ["", "Violencia", "Violencia laboral"]
    ]

    # Llenar datos con puntuaciones reales
    for row_idx, (cat, dominio, dimension) in enumerate(categorias_data, start=8):
        # Celda de categoría
        ws.cell(row=row_idx, column=1, value=cat).border = border
        if cat == "":  # Si está vacío, sombrear
            ws.cell(row=row_idx, column=1).fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Celda de dominio
        ws.cell(row=row_idx, column=2, value=dominio).border = border
        if dominio == "":  # Si está vacío, sombrear
            ws.cell(row=row_idx, column=2).fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Celda de dimensión
        ws.cell(row=row_idx, column=3, value=dimension).border = border

        # Puntuación de dimensión (calcula con el tipo de pregunta y respuesta real)
        preguntas = mapeo_dimensiones.get(dimension, [])
        puntuacion = 0
        respuestas = []
        for p in preguntas:
            respuesta = detalles_preguntas.get(f"P{p}", "")
            respuestas.append(respuesta)
            tipo = 'negativas' if (1 <= p <= 17 or 34 <=
                                   p <= 46) else 'positivas'
            if respuesta:  # Solo suma si hay respuesta
                puntuacion += puntuaciones[tipo].get(respuesta, 0)
        ws.cell(row=row_idx, column=4, value=puntuacion if respuestas and any(
            respuestas) else "").border = border

        # Resultado del cuestionario (respuestas, muestra vacío si no hay)
        ws.cell(row=row_idx, column=5, value=", ".join(
            [r for r in respuestas if r])).border = border

        # Calificación de la categoría (para filas de categoría principal)
        if cat and not dominio and not dimension:
            pass  # Implementar lógica para categorías principales si es necesario

        # Resultado por dominio (para filas de dominio principal)
        if dominio and not dimension:
            pass  # Implementar lógica para dominios principales si es necesario

    # Fórmula de suma total
    ws['D28'] = f"=SUM(D8:D27)"
    ws['D28'].border = border

    # Recomendaciones finales
    ws.merge_cells('A30:G35')
    recomendacion = generar_recomendaciones(row['Nivel de Riesgo'])
    ws['A30'] = f"RECOMENDACIONES:\n\n{recomendacion}"
    ws['A30'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A30'].font = Font(bold=True)

    # Ajustar anchos de columna
    column_widths = [25, 25, 35, 20, 25, 25, 25]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return wb


def main():
    """
    Función principal: pide al usuario seleccionar el archivo de entrada y la carpeta de salida,
    procesa los datos y genera los reportes.
    """
    try:
        # Seleccionar archivo Excel de entrada
        root = tk.Tk()
        root.withdraw()
        archivo_excel = filedialog.askopenfilename(
            title="Selecciona el archivo Excel a evaluar",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")]
        )
        if not archivo_excel:
            print("No se seleccionó ningún archivo.")
            return

        # Seleccionar carpeta de destino para los resultados
        carpeta_destino = filedialog.askdirectory(
            title="Selecciona la carpeta donde se guardarán los resultados"
        )
        if not carpeta_destino:
            print("No se seleccionó ninguna carpeta de destino.")
            return

        # Leer datos de entrada
        df = pd.read_excel(
            archivo_excel, sheet_name='Respuestas de formulario 1')
        df.columns = [col.split('.')[0] if '.' in str(
            col) else col for col in df.columns]

        # Calcular puntuaciones y detalles
        resultados, detalles_preguntas = calcular_puntuaciones(df)

        print("\nResultados de la evaluación de riesgos psicosociales:")
        print(resultados[['Nombre', 'Puntuación Total', 'Nivel de Riesgo']])

        # Guardar archivo general con todos los resultados
        archivo_general = os.path.join(
            carpeta_destino, 'resultados_evaluacion_psicosocial.xlsx')
        resultados.to_excel(archivo_general, index=False)

        # Crear carpeta para archivos individuales dentro de la carpeta destino
        carpeta_individuales = os.path.join(
            carpeta_destino, "resultados_individuales")
        os.makedirs(carpeta_individuales, exist_ok=True)

        # Crear reportes individuales para cada trabajador
        for idx, row in resultados.iterrows():
            nombre = row['Nombre']
            detalles = detalles_preguntas.iloc[idx]

            # Crear reporte individual
            wb = crear_reporte_individual(
                row, detalles, area_adscrita="Área por definir")

            # Guardar archivo
            archivo_individual = os.path.join(
                carpeta_individuales, f"Reporte_{nombre.replace(' ', '_')}.xlsx")
            wb.save(archivo_individual)
            print(f"Reporte creado para: {nombre}")

    except Exception as e:
        print(f"Error al procesar los datos: {str(e)}")


class App:
    """
    Clase principal de la interfaz gráfica.
    Permite seleccionar archivo, carpeta y procesar los reportes.
    """

    def __init__(self, root):
        self.root = root
        self.root.title("Evaluador de Riesgos Psicosociales")
        self.root.geometry("700x500")
        self.archivo_excel = None
        self.carpeta_destino = None

        # Etiquetas y botones de la interfaz
        Label(root, text="Evaluador de Riesgos Psicosociales",
              font=("Arial", 16, "bold")).pack(pady=10)
        self.label_archivo = Label(root, text="Archivo Excel: No seleccionado")
        self.label_archivo.pack(pady=5)
        Button(root, text="Seleccionar archivo Excel",
               command=self.seleccionar_archivo).pack(pady=5)

        self.label_carpeta = Label(
            root, text="Carpeta de destino: No seleccionada")
        self.label_carpeta.pack(pady=5)
        Button(root, text="Seleccionar carpeta de destino",
               command=self.seleccionar_carpeta).pack(pady=5)

        Button(root, text="Procesar y generar reportes",
               command=self.procesar, bg="#4F81BD", fg="white").pack(pady=20)

        # Área de texto para la vista previa del archivo
        self.text_preview = Text(root, height=10, width=80, wrap="none")
        self.text_preview.pack(pady=10)
        self.text_preview.config(state='disabled')
        # Scrollbar opcional
        scrollbar = Scrollbar(root, command=self.text_preview.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.text_preview['yscrollcommand'] = scrollbar.set

    def seleccionar_archivo(self):
        """
        Permite al usuario seleccionar el archivo Excel de entrada y muestra una vista previa.
        """
        archivo = filedialog.askopenfilename(
            title="Selecciona el archivo Excel a evaluar",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")]
        )
        if archivo:
            self.archivo_excel = archivo
            self.label_archivo.config(
                text=f"Archivo Excel: {os.path.basename(archivo)}")
            # Mostrar vista previa
            try:
                df = pd.read_excel(archivo)
                preview = df.head(5).to_string(index=False)
                self.text_preview.config(state='normal')
                self.text_preview.delete(1.0, END)
                self.text_preview.insert(END, preview)
                self.text_preview.config(state='disabled')
            except Exception as e:
                self.text_preview.config(state='normal')
                self.text_preview.delete(1.0, END)
                self.text_preview.insert(
                    END, f"Error al leer el archivo: {str(e)}")
                self.text_preview.config(state='disabled')

    def seleccionar_carpeta(self):
        """
        Permite al usuario seleccionar la carpeta de destino.
        """
        carpeta = filedialog.askdirectory(
            title="Selecciona la carpeta de destino")
        if carpeta:
            self.carpeta_destino = carpeta
            self.label_carpeta.config(text=f"Carpeta de destino: {carpeta}")

    def procesar(self):
        """
        Procesa el archivo seleccionado y genera los reportes.
        """
        if not self.archivo_excel or not self.carpeta_destino:
            messagebox.showerror(
                "Error", "Debes seleccionar el archivo y la carpeta de destino.")
            return
        try:
            df = pd.read_excel(self.archivo_excel,
                               sheet_name='Respuestas de formulario 1')
            df.columns = [col.split('.')[0] if '.' in str(
                col) else col for col in df.columns]
            resultados, detalles_preguntas = calcular_puntuaciones(df)

            archivo_general = os.path.join(
                self.carpeta_destino, 'resultados_evaluacion_psicosocial.xlsx')
            resultados.to_excel(archivo_general, index=False)

            carpeta_individuales = os.path.join(
                self.carpeta_destino, "resultados_individuales")
            os.makedirs(carpeta_individuales, exist_ok=True)

            for idx, row in resultados.iterrows():
                nombre = row['Nombre']
                detalles = detalles_preguntas.iloc[idx]
                wb = crear_reporte_individual(
                    row, detalles, area_adscrita="Área por definir")
                archivo_individual = os.path.join(
                    carpeta_individuales, f"Reporte_{nombre.replace(' ', '_')}.xlsx")
                wb.save(archivo_individual)

            messagebox.showinfo("Éxito", "¡Reportes generados correctamente!")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")


if __name__ == "__main__":
    # Inicia la interfaz gráfica
    root = Tk()
    app = App(root)
    root.mainloop()
